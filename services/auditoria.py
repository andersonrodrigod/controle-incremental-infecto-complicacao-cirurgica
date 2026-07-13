from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.caminhos import obter_caminho_fluxo
from core.config import carregar_configuracoes


FORMATO_DATA_HORA_AUDITORIA = "%Y-%m-%d %H:%M:%S"


def _carregar_ou_criar_planilha(
    caminho_auditoria,
    nome_aba: str
) -> tuple[Workbook, Worksheet]:
    if caminho_auditoria.exists() and caminho_auditoria.stat().st_size > 0:
        workbook = load_workbook(caminho_auditoria)
        if nome_aba in workbook.sheetnames:
            return workbook, workbook[nome_aba]

        return workbook, workbook.create_sheet(nome_aba)

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = nome_aba #type: ignore

    return workbook, planilha #type: ignore


def _obter_cabecalhos(planilha: Worksheet) -> list[str]:
    if planilha.max_row == 0:
        return []

    valores = [
        celula.value
        for celula in planilha[1]
    ]

    return [
        str(valor)
        for valor in valores
        if valor not in (None, "")
    ]


def _garantir_cabecalhos(
    planilha: Worksheet,
    registro: dict[str, Any]
) -> list[str]:
    cabecalhos = _obter_cabecalhos(planilha)

    for coluna in registro:
        if coluna not in cabecalhos:
            cabecalhos.append(coluna)
            planilha.cell(
                row=1,
                column=len(cabecalhos),
                value=coluna
            )

    return cabecalhos


def registrar_auditoria_execucao(
    registro: dict[str, Any],
    nome_fluxo: str
) -> None:
    configuracoes = carregar_configuracoes()
    configuracao_auditoria = configuracoes["fluxos"][nome_fluxo][
        "auditoria"
    ]
    caminho_auditoria = obter_caminho_fluxo(nome_fluxo, "auditoria")

    caminho_auditoria.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    registro_auditoria = {
        "data_hora": datetime.now().strftime(
            FORMATO_DATA_HORA_AUDITORIA
        ),
        **registro
    }

    workbook, planilha = _carregar_ou_criar_planilha(
        caminho_auditoria=caminho_auditoria,
        nome_aba=configuracao_auditoria["aba"]
    )
    cabecalhos = _garantir_cabecalhos(
        planilha=planilha,
        registro=registro_auditoria
    )
    proxima_linha = planilha.max_row + 1

    for indice_coluna, coluna in enumerate(cabecalhos, start=1):
        planilha.cell(
            row=proxima_linha,
            column=indice_coluna,
            value=registro_auditoria.get(coluna, "")
        )

    workbook.save(caminho_auditoria)
