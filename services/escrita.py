from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def salvar_excel(
    dados: pd.DataFrame,
    caminho_destino: str | Path,
    nome_aba: str,
    dados_incrementais: pd.DataFrame | None = None,
) -> None:
    """
    Salva um DataFrame em um arquivo Excel.

    Se o arquivo ja existir e houver dados incrementais, acrescenta somente
    esses registros preservando a formatacao da planilha.
    """

    caminho = Path(caminho_destino)

    caminho.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    if caminho.exists() and dados_incrementais is not None:
        acrescentar_registros_excel_preservando_formato(
            dados=dados_incrementais,
            caminho_destino=caminho,
            nome_aba=nome_aba,
        )
        return

    dados.to_excel(
        caminho,
        sheet_name=nome_aba,
        index=False
    )


def acrescentar_registros_excel_preservando_formato(
    dados: pd.DataFrame,
    caminho_destino: str | Path,
    nome_aba: str,
) -> None:
    """
    Acrescenta registros ao final de uma aba existente sem recriar o arquivo.
    """

    if dados.empty:
        return

    caminho = Path(caminho_destino)
    workbook = load_workbook(caminho)

    if nome_aba not in workbook.sheetnames:
        raise ValueError(
            f"Nao foi possivel localizar a aba '{nome_aba}' "
            f"no arquivo '{caminho.name}'."
        )

    planilha = workbook[nome_aba]
    colunas_planilha = _obter_colunas_cabecalho(planilha)
    colunas_faltantes = [
        coluna
        for coluna in dados.columns
        if coluna not in colunas_planilha
    ]

    if colunas_faltantes:
        _adicionar_colunas_ao_cabecalho(
            planilha=planilha,
            colunas=colunas_faltantes,
        )
        colunas_planilha = _obter_colunas_cabecalho(planilha)

    ultima_linha = planilha.max_row
    linha_modelo = ultima_linha if ultima_linha > 1 else 1

    for _, registro in dados.iterrows():
        ultima_linha += 1
        _copiar_estilo_linha(
            planilha=planilha,
            linha_origem=linha_modelo,
            linha_destino=ultima_linha,
        )

        for indice_coluna, nome_coluna in enumerate(colunas_planilha, start=1):
            if nome_coluna in dados.columns:
                valor = registro[nome_coluna]
                planilha.cell(
                    row=ultima_linha,
                    column=indice_coluna,
                    value=_normalizar_valor_excel(valor),
                )

    workbook.save(caminho)


def _obter_colunas_cabecalho(planilha: Worksheet) -> list[str]:
    return [
        str(celula.value).strip()
        for celula in planilha[1]
        if celula.value is not None and str(celula.value).strip() != ""
    ]


def _adicionar_colunas_ao_cabecalho(
    planilha: Worksheet,
    colunas: list[str],
) -> None:
    ultima_coluna = planilha.max_column
    coluna_modelo = ultima_coluna if ultima_coluna > 0 else 1

    for deslocamento, nome_coluna in enumerate(colunas, start=1):
        destino = planilha.cell(row=1, column=ultima_coluna + deslocamento)
        origem = planilha.cell(row=1, column=coluna_modelo)
        _copiar_estilo_celula(origem=origem, destino=destino)
        destino.value = nome_coluna


def _copiar_estilo_linha(
    planilha: Worksheet,
    linha_origem: int,
    linha_destino: int,
) -> None:
    planilha.row_dimensions[linha_destino].height = (
        planilha.row_dimensions[linha_origem].height
    )

    for indice_coluna in range(1, planilha.max_column + 1):
        origem = planilha.cell(row=linha_origem, column=indice_coluna)
        destino = planilha.cell(row=linha_destino, column=indice_coluna)
        _copiar_estilo_celula(origem=origem, destino=destino)


def _copiar_estilo_celula(origem, destino) -> None:
    if origem.has_style:
        destino.font = copy(origem.font)
        destino.fill = copy(origem.fill)
        destino.border = copy(origem.border)
        destino.alignment = copy(origem.alignment)
        destino.number_format = origem.number_format
        destino.protection = copy(origem.protection)


def _normalizar_valor_excel(valor):
    if pd.isna(valor):
        return None

    return valor
