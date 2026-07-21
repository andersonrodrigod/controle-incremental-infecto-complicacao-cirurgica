import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from services.escrita import salvar_excel


def test_salvar_excel_acrescenta_incrementais_sem_sobrescrever_existentes(
    tmp_path,
):
    caminho = tmp_path / "controle.xlsx"
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "BASE"
    planilha.append(["SENHA", "USUARIO"])
    planilha.append(["100", "Maria manual"])
    planilha["A2"].fill = PatternFill(
        fill_type="solid",
        fgColor="FFFF00",
    )
    workbook.save(caminho)

    dados = pd.DataFrame(
        [
            {"SENHA": "100", "USUARIO": "Maria atualizada"},
            {"SENHA": "101", "USUARIO": "Joao"},
        ]
    )
    dados_incrementais = pd.DataFrame(
        [
            {"SENHA": "101", "USUARIO": "Joao"},
        ]
    )

    salvar_excel(
        dados=dados,
        caminho_destino=caminho,
        nome_aba="BASE",
        dados_incrementais=dados_incrementais,
    )

    workbook_resultado = load_workbook(caminho)
    planilha_resultado = workbook_resultado["BASE"]

    assert planilha_resultado.max_row == 3
    assert planilha_resultado["A2"].value == "100"
    assert planilha_resultado["B2"].value == "Maria manual"
    assert planilha_resultado["A3"].value == "101"
    assert planilha_resultado["B3"].value == "Joao"
    assert planilha_resultado["A3"].fill.fgColor.rgb == "00FFFF00"


def test_salvar_excel_cria_arquivo_quando_destino_nao_existe(tmp_path):
    caminho = tmp_path / "controle.xlsx"
    dados = pd.DataFrame([{"SENHA": "100", "USUARIO": "Maria"}])

    salvar_excel(
        dados=dados,
        caminho_destino=caminho,
        nome_aba="BASE",
        dados_incrementais=None,
    )

    resultado = pd.read_excel(caminho, sheet_name="BASE")

    assert resultado.to_dict("records") == [
        {"SENHA": 100, "USUARIO": "Maria"},
    ]
